import openpyxl
from datetime import datetime
import os 
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

EXCEL_FILE = "logs/ALuL Issue Tracking.xlsx"

def excel_checker ():
    """
    Ensure that the Excel log file and its directory exist.
    Create them if they do not exist.
    """

    Path ("logs").mkdir(exist_ok=True)

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet is not None:
            sheet.title = "Production Tracking"
        
        # Add headers
        headers = ["log_id", "date_reported", "reporter", "issue_type", "equipment_id", "issue_summary", "severity"]

        if sheet is not None:
            sheet.append(headers)

            header_font = Font(name="Apos Narrow", size=9, bold=True)
            header_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Create table
            tab = Table(displayName="IssueTrackingTable", ref="A1:G1")
            style = TableStyleInfo(name="None", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            sheet.add_table(tab)

        wb.save(EXCEL_FILE)

def get_next_log_id(sheet) -> int:
    for r in range(sheet.max_row, 1, -1):
        v = sheet.cell(row=r, column=1).value
        if isinstance(v, int):
            return v + 1
        if isinstance(v, str) and v.strip().isdigit():
            return int(v.strip()) + 1

    return max(1, sheet.max_row)

async def log_report_to_excel(reporter: str, type: str, equipment: str, issue_summary: str, severity: str):
    """Log the report and AI analysis to Excel"""
    try:
        excel_checker()
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active

        next_id = get_next_log_id(sheet)

        row_data = [
            f"LOG_{next_id}",                              # log_id
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # date_reported
            reporter,
            type,
            equipment,
            issue_summary,
            severity
        ]

        if sheet is not None:
            sheet.append(row_data)

            # Update table range to include new row
            current_row = sheet.max_row
            for table in sheet.tables.values():
                table.ref = f"A1:G{current_row}"

            # Define color for each type
            color_map = {
                "DATA_PROB": "FFFF00",
                "HARDWARE": "FF0000",
                "SOFTWARE": "0000FF",
                "MISMATCH": "FFA500",
                "OTHERS": "808080",
                "UNSURE": "FFFFFF"
            }

            fill_color = color_map.get(type.upper(), "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            sheet[f"D{current_row}"].fill = fill

            main_font = Font(name="Apos Narrow", size=9)
            for cell in sheet[current_row]:
                cell.font = main_font
        
        wb.save(EXCEL_FILE)        
        return True
        
    except Exception as e:
        print(f"Error logging to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False